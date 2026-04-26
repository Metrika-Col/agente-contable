"""
Agente Auxiliar Contable — Metrika Group
FastAPI + Twilio WhatsApp + openpyxl
"""
import os, io, re, logging, math, time
from collections import defaultdict
from datetime import datetime, date, timedelta
from typing import Optional

import httpx
import pdfplumber
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, Form, BackgroundTasks, HTTPException, UploadFile, File, Request
from fastapi.responses import JSONResponse, FileResponse
import anthropic
from twilio.rest import Client as TwilioClient

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("agente_contable")

ANTHROPIC_API_KEY  = os.getenv("ANTHROPIC_API_KEY", "")
GEMINI_API_KEY     = os.getenv("GEMINI_API_KEY", "")
TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN  = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_WA_NUMBER   = os.getenv("TWILIO_WA_NUMBER", "whatsapp:+14155238886")

claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
twilio = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

if GEMINI_API_KEY:
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    _gemini = genai.GenerativeModel("gemini-2.0-flash-lite")
else:
    _gemini = None

_GEMINI_SYSTEM = "Eres CONTA, auxiliar contable de Metrika Group. Responde en español, máximo 800 caracteres, conciso y directo."

app = FastAPI(title="Agente Auxiliar Contable", version="1.0.0")

# ─── Sesiones por número de teléfono (TTL 24 h) ──────────────────────────────
from dataclasses import dataclass, field as dc_field

@dataclass
class SesionExtracto:
    mov_banco:    list[dict]
    resumen:      dict
    clasificados: list[dict]
    facturas:     list[dict] = dc_field(default_factory=list)
    timestamp:    datetime   = dc_field(default_factory=datetime.now)

SESIONES: dict[str, SesionExtracto] = {}
SESION_TTL_HORAS = 24

def limpiar_sesiones_expiradas():
    ahora = datetime.now()
    expiradas = [n for n, s in SESIONES.items()
                 if (ahora - s.timestamp).total_seconds() > SESION_TTL_HORAS * 3600]
    for n in expiradas:
        del SESIONES[n]
        log.info(f"Sesión expirada eliminada: {n}")

# ─── Reglas locales de clasificación PUC ─────────────────────────────────────
REGLAS_PUC_LOCAL = [
    (["intereses ahorros", "abono intereses", "intereses"],  "1110", "Bancos"),
    (["nequi", "transferencia nequi"],                        "1305", "Clientes"),
    (["nomina", "nomi ", "salario", "empleado"],              "5105", "Gastos de personal"),
    (["farmatodo", "farmacia", "drogueria", "drogas"],        "5305", "Otros gastos operacionales"),
    (["mercadopago", "mercado pago"],                         "2205", "Proveedores nacionales"),
    (["claude.ai", "subscri", "suscripci"],                   "5305", "Otros gastos operacionales"),
    (["gmf", "4x1000", "gravamen movimiento"],                "5305", "Otros gastos operacionales"),
    (["combustible", "gasolina", "combustibl"],               "5210", "Gastos de viaje"),
    (["retenci", "retencion", "ret. fte"],                    "2365", "Retención en la fuente a pagar"),
    (["iva ", "impuesto ventas"],                             "2408", "IVA generado"),
    (["energia", "electrica", "epm ", "codensa"],             "5115", "Servicios públicos"),
    (["acueducto", "agua ", "triple a"],                      "5115", "Servicios públicos"),
    (["telefon", "celular", "internet", "claro", "movistar"], "5120", "Comunicaciones"),
    (["arrend", "alquiler", "canon"],                         "5135", "Arrendamientos"),
    (["seguro", "prima ", "bolivar", "sura "],                "5165", "Seguros"),
    (["pago nomi", "pago de nomi"],                           "5105", "Gastos de personal"),
    (["transferencia"],                                        "1110", "Bancos"),
]

def clasificar_con_reglas_locales(mov_banco: list[dict]) -> list[dict]:
    resultado = []
    for m in mov_banco:
        desc = m.get("concepto", "").lower()
        cuenta, nombre = None, None
        for keywords, cod, nom in REGLAS_PUC_LOCAL:
            if any(k in desc for k in keywords):
                cuenta, nombre = cod, nom
                break
        resultado.append({**m, "cuenta_puc": cuenta, "nombre_puc": nombre})
    return resultado

def computar_resumen_wa(mov_banco: list[dict]) -> dict:
    ingresos = sum(m["credito"] for m in mov_banco)
    egresos  = sum(m["debito"]  for m in mov_banco)
    top10 = sorted(mov_banco, key=lambda m: m["credito"] + m["debito"], reverse=True)[:10]
    dist: dict[str, dict] = {}
    for m in mov_banco:
        cod = m.get("cuenta_puc") or "Sin clasificar"
        nom = m.get("nombre_puc") or "Sin clasificar"
        if cod not in dist:
            dist[cod] = {"nombre": nom, "n": 0, "total": 0.0}
        dist[cod]["n"] += 1
        dist[cod]["total"] += m["credito"] + m["debito"]
    return {
        "total_tx":   len(mov_banco),
        "ingresos":   ingresos,
        "egresos":    egresos,
        "saldo":      ingresos - egresos,
        "top10":      top10,
        "por_cuenta": dist,
    }

def fmt_cop(n: float) -> str:
    return f"${n:,.0f}".replace(",", ".")

# ─── Parsers de extracto Bancolombia ─────────────────────────────────────────

async def parsear_extracto_bancolombia(pdf_bytes: bytes) -> list[dict]:
    log.info(f"Iniciando parser — PDF recibido: {len(pdf_bytes):,} bytes")

    log.info("Parser nivel 1: texto plano + regex Bancolombia ahorros...")
    try:
        movimientos = _parsear_bancolombia_ahorros(pdf_bytes)
        if len(movimientos) > 10:
            log.info(f"Parser nivel 1 OK — {len(movimientos)} movimientos extraídos")
            return movimientos
        log.info(f"Parser nivel 1: {len(movimientos)} mov (umbral 10), pasando al nivel 2")
    except Exception as e:
        log.error(f"Parser nivel 1 FALLÓ — {type(e).__name__}: {e}", exc_info=True)

    log.info("Parser nivel 2: pdfplumber tablas (fecha completa)...")
    try:
        movimientos = _parsear_con_tablas(pdf_bytes)
        if movimientos:
            log.info(f"Parser nivel 2 OK — {len(movimientos)} movimientos extraídos")
            return movimientos
        log.info("Parser nivel 2: sin resultados — PDF no compatible")
    except Exception as e:
        log.error(f"Parser nivel 2 FALLÓ — {type(e).__name__}: {e}", exc_info=True)

    log.error("Ambos parsers fallaron o devolvieron 0 movimientos")
    return []


# Formato Bancolombia Ahorros: D/MM  DESCRIPCION  VALOR  SALDO
_RX_BANC_TX = re.compile(
    r"^(\d{1,2}/\d{2})"
    r"\s+(.+?)"
    r"\s+(-?[\d,]*\.\d{2})"
    r"\s+([\d,]*\.\d{2})\s*$"
)
_RX_DESDE = re.compile(r"DESDE:\s*(\d{4})/(\d{2})/\d{2}")


def _parsear_bancolombia_ahorros(pdf_bytes: bytes) -> list[dict]:
    movimientos: list[dict] = []
    desde_year  = datetime.now().year
    desde_month = 1

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages_text: list[str] = []
        all_text = ""
        for page in pdf.pages:
            t = page.extract_text() or ""
            pages_text.append(t)
            all_text += t + "\n"

        m = _RX_DESDE.search(all_text)
        if m:
            desde_year  = int(m.group(1))
            desde_month = int(m.group(2))

        current_year = desde_year
        prev_month   = desde_month

        for text in pages_text:
            for line in text.split("\n"):
                line = line.strip()
                mx = _RX_BANC_TX.match(line)
                if not mx:
                    continue

                fecha_str, concepto, valor_str, saldo_str = mx.groups()
                day_s, month_s = fecha_str.split("/")
                tx_month = int(month_s)
                tx_day   = int(day_s)

                if tx_month < prev_month:
                    current_year += 1
                prev_month = tx_month

                try:
                    fecha   = datetime(current_year, tx_month, tx_day).date()
                    valor   = _parse_valor_banc(valor_str)
                    saldo   = _parse_valor_banc(saldo_str)
                    debito  = abs(valor) if valor < 0 else 0.0
                    credito = valor      if valor > 0 else 0.0
                    movimientos.append({
                        "fecha":    str(fecha),
                        "doc":      "",
                        "concepto": concepto.strip(),
                        "debito":   debito,
                        "credito":  credito,
                        "saldo":    saldo,
                    })
                except (ValueError, Exception):
                    continue

    return movimientos


def _parsear_con_tablas(pdf_bytes: bytes) -> list[dict]:
    movimientos: list[dict] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in (table or []):
                    if not row or len(row) < 4:
                        continue
                    fecha_str = str(row[0]).strip() if row[0] else ""
                    if not re.match(r"\d{2}/\d{2}/\d{4}", fecha_str):
                        continue
                    try:
                        fecha = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                        movimientos.append({
                            "fecha":    str(fecha),
                            "doc":      str(row[1]).strip() if row[1] else "",
                            "concepto": str(row[2]).strip() if row[2] else "",
                            "debito":   _parse_valor_co(str(row[3]) if row[3] else ""),
                            "credito":  _parse_valor_co(str(row[4]) if len(row) > 4 and row[4] else ""),
                            "saldo":    _parse_valor_co(str(row[5]) if len(row) > 5 and row[5] else ""),
                        })
                    except Exception:
                        continue
    return movimientos


def _parse_valor_banc(s: str) -> float:
    s = s.strip().replace("$", "").replace(",", "").replace(" ", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def _parse_valor_co(s: str) -> float:
    s = s.strip().replace("$", "").replace(".", "").replace(",", ".").replace(" ", "")
    try:
        return float(s)
    except Exception:
        return 0.0

# ─── Conciliación: solo datos reales del banco ───────────────────────────────

def conciliar(mov_banco: list[dict]) -> dict:
    if mov_banco and "cuenta_puc" not in mov_banco[0]:
        mov_banco = clasificar_con_reglas_locales(mov_banco)

    clasificados   = [m for m in mov_banco if m.get("cuenta_puc")]
    sin_clasificar = [m for m in mov_banco if not m.get("cuenta_puc")]

    top_egresos = sorted(
        [m for m in mov_banco if m["debito"] > 0],
        key=lambda m: m["debito"],
        reverse=True
    )[:20]

    total_db = sum(m["debito"]  for m in mov_banco)
    total_cr = sum(m["credito"] for m in mov_banco)

    return {
        "mov_banco":      mov_banco,
        "clasificados":   clasificados,
        "sin_clasificar": sin_clasificar,
        "top_egresos":    top_egresos,
        "conciliados": clasificados,
        "solo_banco":  sin_clasificar,
        "solo_conta":  [],
        "resumen": {
            "total_tx":             len(mov_banco),
            "total_clasificados":   len(clasificados),
            "total_sin_clasificar": len(sin_clasificar),
            "debitos_banco":        total_db,
            "creditos_banco":       total_cr,
            "saldo_neto":           total_cr - total_db,
            "total_conciliados":    len(clasificados),
            "total_solo_banco":     len(sin_clasificar),
            "total_solo_conta":     0,
            "diferencia_debitos":   0,
            "diferencia_creditos":  0,
        },
        "nota": "Para conciliación completa necesitas importar tus registros contables de OSSADO.",
    }

# ─── Calendario DIAN real (calculado dinámicamente) ──────────────────────────

_MESES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

def _calendario_dian_proximas() -> list[dict]:
    hoy = date.today()
    obligaciones: list[dict] = []

    for delta in range(4):
        mes_venc = hoy.month + delta
        anio_venc = hoy.year + (mes_venc - 1) // 12
        mes_venc  = ((mes_venc - 1) % 12) + 1
        try:
            venc = date(anio_venc, mes_venc, 12)
        except ValueError:
            continue
        if venc < hoy:
            continue
        mes_per  = (mes_venc - 2) % 12 + 1
        anio_per = anio_venc if mes_venc > 1 else anio_venc - 1
        obligaciones.append({
            "obligacion": "Retención en la fuente",
            "periodo":    f"{_MESES[mes_per]} {anio_per}",
            "vencimiento": str(venc),
        })

    for m1, m2, m_venc in [(1,2,3),(3,4,5),(5,6,7),(7,8,9),(9,10,11),(11,12,1)]:
        anio_venc = hoy.year + (1 if m_venc < m1 else 0)
        try:
            venc = date(anio_venc, m_venc, 12)
        except ValueError:
            continue
        if venc < hoy:
            continue
        anio_per = anio_venc - (1 if m_venc < m1 else 0)
        obligaciones.append({
            "obligacion": "IVA bimestral",
            "periodo":    f"{_MESES[m1]}-{_MESES[m2]} {anio_per}",
            "vencimiento": str(venc),
        })

    for m_ini, m_fin, m_venc in [(1,3,4),(4,6,7),(7,9,10),(10,12,1)]:
        anio_venc = hoy.year + (1 if m_venc < m_ini else 0)
        try:
            venc = date(anio_venc, m_venc, 20)
        except ValueError:
            continue
        if venc < hoy:
            continue
        anio_per = anio_venc - (1 if m_venc < m_ini else 0)
        obligaciones.append({
            "obligacion": "ICA Barranquilla",
            "periodo":    f"T {_MESES[m_ini]}-{_MESES[m_fin]} {anio_per}",
            "vencimiento": str(venc),
        })

    obligaciones.sort(key=lambda o: o["vencimiento"])
    return obligaciones[:7]

# ─── Generación Excel ─────────────────────────────────────────────────────────

def generar_excel_conciliacion(resultado: dict) -> bytes:
    wb = openpyxl.Workbook()
    AMARILLO  = "FFCC00"
    NEGRO     = "1A1A1A"
    VERDE_OSC = "1E6B3C"
    ROJO_OSC  = "9B1C1C"
    AZUL_OSC  = "1E3A5F"
    GRIS_CLARO  = "F5F5F5"
    VERDE_CLA   = "D4EDDA"
    ROJO_CLA    = "F8D7DA"
    AMARILLO_CL = "FFF3CD"

    def hdr_fill(color): return PatternFill("solid", fgColor=color)
    def font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def border_thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def fmt_col(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def escribir_encabezado(ws, titulo, subtitulo):
        ws.merge_cells("A1:H1")
        ws["A1"] = "EXTRACTO BANCOLOMBIA — ANÁLISIS CONTABLE"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        ws["A1"].fill = hdr_fill(NEGRO)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A2:H2")
        ws["A2"] = titulo
        ws["A2"].font = Font(bold=True, size=12, color=NEGRO, name="Calibri")
        ws["A2"].fill = hdr_fill(AMARILLO)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 24
        ws.merge_cells("A3:H3")
        ws["A3"] = subtitulo
        ws["A3"].font = Font(size=9, color="555555", name="Calibri")
        ws["A3"].alignment = Alignment(horizontal="center")

    mov_banco    = resultado.get("mov_banco", [])
    clasificados = resultado.get("clasificados", [])
    sin_clasi    = resultado.get("sin_clasificar", [])
    top_egresos  = resultado.get("top_egresos", [])
    res          = resultado["resumen"]

    ws1 = wb.active
    ws1.title = "Clasificados PUC"
    escribir_encabezado(ws1, "MOVIMIENTOS CLASIFICADOS POR PUC",
                        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    cols1 = ["Fecha", "Concepto", "Débito ($)", "Crédito ($)", "Saldo ($)", "Cuenta PUC", "Nombre PUC"]
    for j, c in enumerate(cols1, 1):
        cell = ws1.cell(row=5, column=j, value=c)
        cell.font  = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill  = hdr_fill(VERDE_OSC)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
    ws1.row_dimensions[5].height = 28
    for i, m in enumerate(clasificados, 6):
        concepto = m["concepto"]
        vals = [m["fecha"], concepto, m["debito"] or None,
                m["credito"] or None, m.get("saldo") or None,
                m.get("cuenta_puc", ""), m.get("nombre_puc", "")]
        for j, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=j, value=v)
            cell.font   = font(size=9)
            cell.border = border_thin()
            cell.fill   = hdr_fill(VERDE_CLA) if i % 2 == 0 else hdr_fill("FFFFFF")
            if j == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif j in (3, 4, 5) and v:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        descripcion = str(ws1.cell(row=i, column=2).value or "")
        lineas = math.ceil(len(descripcion) / 45)
        ws1.row_dimensions[i].height = max(18, lineas * 15)
    widths1 = [12, 45, 14, 14, 14, 12, 28]
    for j, w in enumerate(widths1, 1): fmt_col(ws1, j, w)

    ws2 = wb.create_sheet("Sin Clasificar")
    escribir_encabezado(ws2, "MOVIMIENTOS SIN CUENTA PUC",
                        "Requieren revisión manual — asignar cuenta contable")
    ws2.merge_cells("A4:G4")
    nota = ws2.cell(row=4, column=1,
        value="⚠️  Para conciliación completa, importar registros contables de OSSADO")
    nota.font = Font(bold=True, color=ROJO_OSC, size=10, name="Calibri")
    nota.alignment = Alignment(horizontal="center")
    cols2 = ["Fecha", "Concepto", "Débito ($)", "Crédito ($)", "Saldo ($)", "Tipo", "Acción"]
    for j, c in enumerate(cols2, 1):
        cell = ws2.cell(row=5, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill = hdr_fill(ROJO_OSC)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin()
    for i, m in enumerate(sin_clasi, 6):
        tipo = "Crédito" if m["credito"] > 0 else "Débito"
        vals = [m["fecha"], m["concepto"], m["debito"] or None,
                m["credito"] or None, m.get("saldo") or None,
                tipo, "Asignar cuenta PUC manualmente"]
        for j, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=j, value=v)
            cell.font = font(size=9)
            cell.border = border_thin()
            cell.fill = hdr_fill(AMARILLO_CL)
            if j == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif j in (3, 4, 5) and v:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        descripcion = str(ws2.cell(row=i, column=2).value or "")
        lineas = math.ceil(len(descripcion) / 45)
        ws2.row_dimensions[i].height = max(18, lineas * 15)
    widths2 = [12, 45, 14, 14, 14, 10, 35]
    for j, w in enumerate(widths2, 1): fmt_col(ws2, j, w)

    ws3 = wb.create_sheet("Top Egresos")
    escribir_encabezado(ws3, "TOP EGRESOS DEL PERÍODO",
                        f"Corte: {date.today().strftime('%d/%m/%Y')}")
    cols3 = ["Fecha", "Concepto", "Débito ($)", "Cuenta PUC", "Nombre PUC"]
    for j, c in enumerate(cols3, 1):
        cell = ws3.cell(row=5, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill = hdr_fill(NEGRO)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin()
    for i, m in enumerate(top_egresos, 6):
        vals = [m["fecha"], m["concepto"], m["debito"],
                m.get("cuenta_puc", "—"), m.get("nombre_puc", "Sin clasificar")]
        for j, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=j, value=v)
            cell.font = font(size=9)
            cell.border = border_thin()
            cell.fill = hdr_fill(ROJO_CLA) if i % 2 == 0 else hdr_fill("FFFFFF")
            if j == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif j == 3:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        descripcion = str(ws3.cell(row=i, column=2).value or "")
        lineas = math.ceil(len(descripcion) / 45)
        ws3.row_dimensions[i].height = max(18, lineas * 15)
    widths3 = [12, 45, 14, 12, 28]
    for j, w in enumerate(widths3, 1): fmt_col(ws3, j, w)
    last3 = 5 + len(top_egresos)
    total_egresos_real = sum((m.get("debito") or 0) for m in clasificados + sin_clasi)
    lbl3 = ws3.cell(row=last3+1, column=2, value="TOTAL EGRESOS (todos los cargos):")
    lbl3.font = Font(bold=True, size=9, name="Calibri")
    lbl3.alignment = Alignment(horizontal="right")
    tc3 = ws3.cell(row=last3+1, column=3, value=total_egresos_real)
    tc3.number_format = "#,##0"
    tc3.font = Font(bold=True, size=10, color=ROJO_OSC, name="Calibri")
    tc3.alignment = Alignment(horizontal="right")
    tc3.border = border_thin()

    ws4 = wb.create_sheet("Calendario DIAN")
    escribir_encabezado(ws4, "PRÓXIMAS OBLIGACIONES FISCALES",
                        f"Calculado desde {date.today().strftime('%d/%m/%Y')} — verificar NIT en DIAN")
    cols4 = ["Obligación", "Período", "Fecha Vencimiento", "Días Restantes", "Estado"]
    for j, c in enumerate(cols4, 1):
        cell = ws4.cell(row=5, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill = hdr_fill(AZUL_OSC)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin()
    hoy = date.today()
    for i, ob in enumerate(_calendario_dian_proximas(), 6):
        venc = datetime.strptime(ob["vencimiento"], "%Y-%m-%d").date()
        dias = (venc - hoy).days
        color = ROJO_CLA if dias < 5 else (AMARILLO_CL if dias <= 15 else VERDE_CLA)
        vals = [ob["obligacion"], ob["periodo"], ob["vencimiento"], dias, "PENDIENTE"]
        for j, v in enumerate(vals, 1):
            cell = ws4.cell(row=i, column=j, value=v)
            cell.font = font(size=9)
            cell.border = border_thin()
            cell.fill = hdr_fill(color)
            if j in (3, 4):
                cell.alignment = Alignment(horizontal="center")
    widths4 = [28, 20, 18, 15, 12]
    for j, w in enumerate(widths4, 1): fmt_col(ws4, j, w)
    last4 = 5 + len(_calendario_dian_proximas())
    nota4 = ws4.cell(row=last4+2, column=1,
        value="* Fechas basadas en calendario general colombiano. Confirmar según dígito verificador del NIT.")
    nota4.font = Font(italic=True, size=8, color="888888", name="Calibri")
    ws4.merge_cells(f"A{last4+2}:E{last4+2}")

    ws5 = wb.create_sheet("Resumen Ejecutivo")
    ws5.sheet_view.showGridLines = False
    escribir_encabezado(ws5, "RESUMEN EJECUTIVO",
                        f"Metrika Group | {date.today().strftime('%d/%m/%Y')}")
    metricas = [
        ("", ""),
        ("MOVIMIENTOS DEL EXTRACTO", ""),
        ("Total transacciones",          res["total_tx"]),
        ("Clasificados con PUC",         res["total_clasificados"]),
        ("Sin clasificar",               res["total_sin_clasificar"]),
        ("Total débitos",                res["debitos_banco"]),
        ("Total créditos",               res["creditos_banco"]),
        ("Saldo neto del período",       res["saldo_neto"]),
        ("", ""),
        ("NOTA", "Para conciliación completa importar registros de OSSADO"),
    ]
    for i, (etiqueta, valor) in enumerate(metricas, 5):
        cell_e = ws5.cell(row=i, column=2, value=etiqueta)
        cell_v = ws5.cell(row=i, column=4, value=valor if valor != "" else "")
        if etiqueta in ("MOVIMIENTOS DEL EXTRACTO",):
            cell_e.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell_e.fill = hdr_fill(NEGRO)
            ws5.merge_cells(f"B{i}:E{i}")
            ws5.row_dimensions[i].height = 22
        elif etiqueta == "NOTA":
            cell_e.font = Font(bold=True, color=AZUL_OSC, size=9, name="Calibri")
            cell_e.fill = hdr_fill(AMARILLO_CL)
            cell_v.font = Font(italic=True, color=AZUL_OSC, size=9, name="Calibri")
            cell_v.fill = hdr_fill(AMARILLO_CL)
            cell_e.border = border_thin()
            cell_v.border = border_thin()
        elif etiqueta:
            cell_e.font = font(size=10)
            cell_e.fill = hdr_fill(GRIS_CLARO)
            if isinstance(valor, (int, float)) and valor > 100:
                cell_v.number_format = "#,##0"
                cell_v.alignment = Alignment(horizontal="right")
            cell_e.border = border_thin()
            cell_v.border = border_thin()
    fmt_col(ws5, 2, 35)
    fmt_col(ws5, 3, 5)
    fmt_col(ws5, 4, 18)

    ws6 = wb.create_sheet("Movimientos")
    escribir_encabezado(ws6, "EXTRACTO COMPLETO — TODOS LOS MOVIMIENTOS",
                        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    AMARILLO_TX = "FFCC00"
    cols6 = ["Fecha", "Descripción", "Valor ($)", "Saldo ($)"]
    for j, c in enumerate(cols6, 1):
        cell = ws6.cell(row=5, column=j, value=c)
        cell.font  = Font(bold=True, color=AMARILLO_TX, size=9, name="Calibri")
        cell.fill  = hdr_fill(NEGRO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
    ws6.row_dimensions[5].height = 28
    todos_movs = sorted(
        clasificados + sin_clasi,
        key=lambda m: m.get("fecha") or ""
    )
    total_abonos = 0.0
    total_cargos = 0.0
    for i, m in enumerate(todos_movs, 6):
        credito = m.get("credito") or 0
        debito  = m.get("debito")  or 0
        valor   = credito if credito > 0 else -debito
        saldo   = m.get("saldo") or None
        if valor > 0:
            total_abonos += valor
        else:
            total_cargos += abs(valor)
        vals6 = [m.get("fecha", ""), m.get("concepto", ""), valor, saldo]
        for j, v in enumerate(vals6, 1):
            cell = ws6.cell(row=i, column=j, value=v)
            cell.font   = font(size=9)
            cell.border = border_thin()
            cell.fill   = hdr_fill(VERDE_CLA) if valor > 0 else hdr_fill(ROJO_CLA)
            if j == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif j in (3, 4) and v is not None:
                cell.number_format = '#,##0;-#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        descripcion6 = str(ws6.cell(row=i, column=2).value or "")
        lineas6 = math.ceil(len(descripcion6) / 45)
        ws6.row_dimensions[i].height = max(18, lineas6 * 15)
    widths6 = [12, 45, 16, 16]
    for j, w in enumerate(widths6, 1): fmt_col(ws6, j, w)
    saldo_neto = total_abonos - total_cargos
    last6 = 5 + len(todos_movs)
    totales6 = [
        ("TOTAL ABONOS:",  total_abonos,  VERDE_OSC),
        ("TOTAL CARGOS:", -total_cargos,  ROJO_OSC),
        ("SALDO NETO:",    saldo_neto,    AZUL_OSC),
    ]
    for offset, (lbl, val, color) in enumerate(totales6, 1):
        row = last6 + offset
        ws6.row_dimensions[row].height = 20
        lbl_cell = ws6.cell(row=row, column=2, value=lbl)
        lbl_cell.font = Font(bold=True, size=9, color=color, name="Calibri")
        lbl_cell.alignment = Alignment(horizontal="right")
        val_cell = ws6.cell(row=row, column=3, value=val)
        val_cell.number_format = '#,##0;-#,##0'
        val_cell.font = Font(bold=True, size=10, color=color, name="Calibri")
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = border_thin()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── Detección y extracción de facturas ───────────────────────────────────────

def _extraer_texto_pdf(pdf_bytes: bytes) -> str:
    texto = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texto += (page.extract_text() or "") + "\n"
    return texto

def _es_factura(texto: str) -> bool:
    t = texto.lower()
    if "factura" in t or "invoice" in t:
        return True
    tiene_nit      = "nit" in t
    tiene_subtotal = "subtotal" in t
    tiene_iva      = bool(re.search(r'\biva\b', t))
    return tiene_nit and (tiene_subtotal or tiene_iva)

def _parse_valor_factura(s: str) -> float:
    """Parse Colombian invoice monetary value.
    Handles: 110.642 · 110,642 · 1.110.642 · 110.642,00 · 110,642.00
    """
    s = s.strip().replace("$", "").replace("\xa0", "").replace(" ", "")
    if not s:
        return 0.0
    ndots   = s.count(".")
    ncommas = s.count(",")
    if ndots > 1:                               # 1.110.642[,00]
        s = s.replace(".", "").replace(",", ".")
    elif ncommas > 1:                           # 1,110,642
        s = s.replace(",", "")
    elif ndots == 1 and ncommas == 1:
        if s.index(".") < s.index(","):         # 110.642,00 → col
            s = s.replace(".", "").replace(",", ".")
        else:                                   # 110,642.00 → US
            s = s.replace(",", "")
    elif ndots == 1:
        if len(s.split(".")[1]) == 3:           # 110.642 → miles
            s = s.replace(".", "")
        # else: decimal → leave as is
    elif ncommas == 1:
        if len(s.split(",")[1]) == 3:           # 110,642 → miles
            s = s.replace(",", "")
        else:                                   # 110,64 → decimal
            s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


# Etiquetas de datos del receptor/cliente que no son el emisor
_RX_ETIQUETA_RECEPTOR = re.compile(
    r'^(Nom|Tele|Tel|Dire|Dir|Ciud|Iden|Fact|Correo|Email|Ciudad|Barrio|Dpto)\s*:',
    re.IGNORECASE
)


def _extraer_datos_factura(texto: str) -> dict:
    t = texto

    # ── EMISOR (cascada) ─────────────────────────────────────────────────────
    emisor = None
    # 1. Campo explícito de factura electrónica
    m = re.search(r"Nombre\s+Raz[oó]n\s+Social\s+Emisor[:\s]*(.*?)(?:[,\n])", t, re.IGNORECASE)
    if m:
        emisor = m.group(1).strip()
    # 2. Línea que contiene "NIT", excluyendo filas de datos del receptor
    if not emisor:
        for linea in t.split('\n'):
            ls = linea.strip()
            if not ls or _RX_ETIQUETA_RECEPTOR.match(ls):
                continue
            if re.search(r'\bNIT\b', ls, re.IGNORECASE):
                mx = re.match(r'^(.+?)\s+NIT', ls, re.IGNORECASE)
                if mx:
                    emisor = mx.group(1).strip()
                    break
    # 3-4. Campos estructurados
    if not emisor:
        m = re.search(r"Raz[oó]n\s+Social[:\s]*(.*?)(?:\n|$)", t, re.IGNORECASE)
        if m: emisor = m.group(1).strip()
    if not emisor:
        m = re.search(r"Empresa[:\s]*(.*?)(?:\n|$)", t, re.IGNORECASE)
        if m: emisor = m.group(1).strip()
    # 5. Fallback: primera línea no vacía que no sea etiqueta de receptor
    if not emisor:
        for linea in t.split('\n'):
            ls = linea.strip()
            if ls and not _RX_ETIQUETA_RECEPTOR.match(ls):
                emisor = ls
                break
    emisor = (emisor or "No detectado")[:50]

    # ── NIT (cascada — guión obligatorio para descartar teléfonos) ───────────
    nit = None
    # 1. "NIT NNN.NNN.NNN-N" — formato con guión requerido
    m = re.search(r'\bNIT\s+(\d{3}\.?\d{3}\.?\d{3}-\d)', t, re.IGNORECASE)
    if m: nit = m.group(1)
    # 2. "NIT: NNN.NNN.NNN-N"
    if not nit:
        m = re.search(r'\bNIT[:\s]+(\d{3}\.?\d{3}\.?\d{3}-\d)', t, re.IGNORECASE)
        if m: nit = m.group(1)
    # 3. Formato con puntos explícitos NNN.NNN.NNN-N en cualquier parte
    if not nit:
        m = re.search(r'\b(\d{3}\.\d{3}\.\d{3}-\d)\b', t)
        if m: nit = m.group(1)
    # 4. "Nit o CC.:" — puede ser número de cliente, último recurso
    if not nit:
        m = re.search(r"Nit\s+o\s+CC\.?:?\s*([\d\.]+)", t, re.IGNORECASE)
        if m: nit = m.group(1).strip()
    nit = nit or "No detectado"

    # ── FECHA (cascada, preferir vencimiento) ────────────────────────────────
    fecha = None
    for pat in [
        r"[Vv]encimiento.*?(\d{1,2}/\w+/\d{4})",
        r"[Vv]encimiento.*?(\d{2}/\d{2}/\d{4})",
        r"[Ff]echa.*?(\d{2}/\d{2}/\d{4})",
        r"(\d{2}-\d{2}-\d{4})",
        r"(\d{2}/\d{2}/\d{4})",
    ]:
        m = re.search(pat, t, re.IGNORECASE)
        if m:
            fecha = m.group(1)
            break
    fecha = fecha or str(date.today())

    # ── TOTAL (cascada + _parse_valor_factura) ───────────────────────────────
    _NUM = r"(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?)"
    total = 0.0
    for pat in [
        rf"[Tt]otal\s+a\s+pagar[^\d\n]*{_NUM}",
        rf"TOTAL\s+A\s+PAGAR[^\d\n]*{_NUM}",
        rf"[Tt]otal\s+a\s+cancelar[^\d\n]*{_NUM}",
        rf"[Vv]alor\s+total[^\d\n]*{_NUM}",
        rf"[Tt]otal\s+factura[^\d\n]*{_NUM}",
        rf"\bTOTAL\b[^\d\n]*{_NUM}",
    ]:
        m = re.search(pat, t, re.IGNORECASE)
        if m:
            v = _parse_valor_factura(m.group(1))
            if v > 0:
                total = v
                break
    # Fallback: mayor valor precedido por $ en el documento
    if total == 0.0:
        candidatos = re.findall(rf'\$\s*{_NUM}', t)
        valores = [_parse_valor_factura(c) for c in candidatos if _parse_valor_factura(c) > 0]
        if valores:
            total = max(valores)

    # ── IVA ──────────────────────────────────────────────────────────────────
    iva = 0.0
    for pat in [
        rf"\bIVA\b[^\d\n]*{_NUM}",
        rf"[Ii]mpuesto[^\d\n]*{_NUM}",
    ]:
        m = re.search(pat, t, re.IGNORECASE)
        if m:
            v = _parse_valor_factura(m.group(1))
            if v > 0:
                iva = v
                break

    # ── SUBTOTAL ─────────────────────────────────────────────────────────────
    subtotal = 0.0
    for pat in [
        rf"[Ss]ubtotal[^\d\n]*{_NUM}",
        rf"[Ss]ub\s+total[^\d\n]*{_NUM}",
    ]:
        m = re.search(pat, t, re.IGNORECASE)
        if m:
            v = _parse_valor_factura(m.group(1))
            if v > 0:
                subtotal = v
                break
    if subtotal == 0.0:
        subtotal = max(0.0, total - iva)

    # ── TIPO Y CUENTA PUC SUGERIDA ───────────────────────────────────────────
    tl = t.lower()
    if any(k in tl for k in ["electricidad", "energia", "eléctrica", "electrica",
                               "epm", "essa", "codensa", "electrificadora"]):
        tipo = "Servicio público - Energía"
        puc_cod, puc_nom = "5115", "Servicios públicos"
    elif any(k in tl for k in ["agua", "acueducto", "alcantarillado", "aaa", "eaab",
                                "triple a", "empresas municipales"]):
        tipo = "Servicio público - Agua"
        puc_cod, puc_nom = "5115", "Servicios públicos"
    elif any(k in tl for k in ["gas natural", "vanti", "surtigas", "gases de occidente",
                                "gases del caribe", "alcanos"]):
        tipo = "Servicio público - Gas"
        puc_cod, puc_nom = "5115", "Servicios públicos"
    elif any(k in tl for k in ["internet", "fibra", "movistar", "claro", "tigo",
                                "etb", "une ", "telecomunicaciones", "telefonia"]):
        tipo = "Telecomunicaciones"
        puc_cod, puc_nom = "5115", "Servicios públicos"
    elif any(k in tl for k in ["arrendamiento", "arriendo", "canon de arriendo",
                                "alquiler"]):
        tipo = "Arrendamiento"
        puc_cod, puc_nom = "5210", "Arrendamientos"
    elif any(k in tl for k in ["honorarios", "consultoría", "consultoria",
                                "servicios profesionales", "asesoría", "asesoria"]):
        tipo = "Servicios profesionales"
        puc_cod, puc_nom = "5305", "Honorarios"
    elif any(k in tl for k in ["factura electrónica de venta", "factura electronica de venta",
                                "supermercado", "almacén", "almacen", "tienda"]):
        tipo = "Compra de bienes"
        puc_cod, puc_nom = "5120", "Gastos generales"
    else:
        tipo = "Factura general"
        puc_cod, puc_nom = "5290", "Otros gastos"

    return {
        "emisor":       emisor,
        "nit":          nit,
        "fecha":        fecha,
        "subtotal":     subtotal,
        "iva":          iva,
        "total":        total,
        "tipo_factura": tipo,
        "cuenta_puc":   puc_cod,
        "nombre_puc":   puc_nom,
        "timestamp":    str(datetime.now().date()),
    }

# ─── WhatsApp helpers ─────────────────────────────────────────────────────────

_WA_MAX  = 1500
_WA_TRUNC = 1450

def _partir_mensaje(texto: str, max_chars: int = _WA_MAX) -> list[str]:
    if len(texto) <= max_chars:
        return [texto]
    partes = []
    while texto:
        if len(texto) <= max_chars:
            partes.append(texto)
            break
        corte = texto.rfind(" ", 0, max_chars)
        if corte == -1:
            corte = max_chars
        partes.append(texto[:corte].rstrip())
        texto = texto[corte:].lstrip()
    return partes

def enviar_whatsapp(to: str, body: str):
    dest = f"whatsapp:{to}" if not to.startswith("whatsapp:") else to
    for parte in _partir_mensaje(body):
        twilio.messages.create(from_=TWILIO_WA_NUMBER, to=dest, body=parte)

def enviar_whatsapp_media(to: str, body: str, media_url: str):
    twilio.messages.create(
        from_=TWILIO_WA_NUMBER,
        to=f"whatsapp:{to}" if not to.startswith("whatsapp:") else to,
        body=body,
        media_url=[media_url]
    )

def subir_excel_twilio(excel_bytes: bytes, filename: str) -> str:
    path = f"/tmp/{filename}"
    with open(path, "wb") as f:
        f.write(excel_bytes)
    size = os.path.getsize(path)
    if size == 0:
        raise RuntimeError(f"El archivo {filename} se escribió vacío en disco")
    log.info(f"Excel guardado en disco: {path} — {size:,} bytes")
    raw_domain = os.getenv("RAILWAY_PUBLIC_DOMAIN", "localhost:8000")
    if raw_domain.startswith("http"):
        base = raw_domain.rstrip("/")
    else:
        base = f"https://{raw_domain.rstrip('/')}"
    url = f"{base}/descargar/{filename}"
    log.info(f"URL de descarga generada: {url}")
    return url

# ─── Sistema de respuestas sin IA ────────────────────────────────────────────

_MESES_NOMBRES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}

_MSG_IDENTIDAD = (
    "🤖 Soy *CONTA*, tu Auxiliar Contable IA de Metrika Group.\n\n"
    "Puedo ayudarte con:\n"
    "📄 Procesar tu extracto bancario (envíame el PDF)\n"
    "💰 Consultar saldos por período\n"
    "📊 Ver ingresos y egresos\n"
    "🔍 Detectar movimientos inusuales\n"
    "🧾 Procesar facturas\n"
    "📅 Calendario fiscal DIAN\n\n"
    "Comandos rápidos:\n"
    "- *saldo* → resumen general\n"
    "- *clasificar* → top 10 con PUC\n"
    "- *impuestos* → fechas DIAN\n"
    "- *pagos pendientes* → principales egresos"
)

_MSG_SIN_EXTRACTO = (
    "📄 Primero envíame tu extracto bancario en PDF para "
    "poder responder esa consulta."
)

def _buscar_patron(mensaje: str, sesion: "SesionExtracto | None") -> "Optional[str]":
    msg = mensaje.lower()

    # 1. Identidad
    if any(k in msg for k in ["qué eres", "que eres", "quién eres", "quien eres",
                               "qué puedes", "que puedes", "ayuda", "help",
                               "comandos", "cómo funciona", "como funciona",
                               "hola", "buenas", "buenos", "menu", "menú"]):
        return _MSG_IDENTIDAD

    # Detectar mes mencionado
    mes_detectado = None
    for nombre_mes, num_mes in _MESES_NOMBRES.items():
        if nombre_mes in msg:
            mes_detectado = (nombre_mes, num_mes)
            break

    # 2. Saldo por mes
    if mes_detectado and any(k in msg for k in ["saldo", "balance", "resumen"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        nombre_mes, num_mes = mes_detectado
        movs_mes = [m for m in sesion.mov_banco
                    if re.match(rf"\d{{4}}-{num_mes:02d}-", m.get("fecha", ""))]
        if not movs_mes:
            return f"ℹ️ No encontré movimientos de {nombre_mes.capitalize()} en el extracto cargado."
        ing_mes = sum(m["credito"] for m in movs_mes)
        eg_mes  = sum(m["debito"]  for m in movs_mes)
        saldo_i = movs_mes[0].get("saldo") or 0
        saldo_f = movs_mes[-1].get("saldo") or 0
        return (
            f"💰 *Saldo de {nombre_mes.capitalize()}*\n\n"
            f"📋 Transacciones: {len(movs_mes)}\n"
            f"💚 Ingresos: {fmt_cop(ing_mes)}\n"
            f"🔴 Egresos: {fmt_cop(eg_mes)}\n"
            f"📈 Saldo neto: {fmt_cop(ing_mes - eg_mes)}\n"
            f"🏦 Saldo inicial: {fmt_cop(saldo_i)}\n"
            f"🏦 Saldo final: {fmt_cop(saldo_f)}"
        )

    # 3. Saldo general
    if "saldo" in msg:
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        r = sesion.resumen
        return (
            f"💰 *Saldo del extracto*\n\n"
            f"💚 Ingresos: {fmt_cop(r['ingresos'])}\n"
            f"🔴 Egresos:  {fmt_cop(r['egresos'])}\n"
            f"💰 Saldo neto: {fmt_cop(r['saldo'])}\n\n"
            f"📋 {r['total_tx']} transacciones en el período"
        )

    # 4. Ingresos
    if any(k in msg for k in ["ingreso", "entró", "entro", "abono", "recibí", "recibi"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        movs = sesion.mov_banco
        if mes_detectado:
            _, num_mes = mes_detectado
            movs = [m for m in movs if re.match(rf"\d{{4}}-{num_mes:02d}-", m.get("fecha", ""))]
        top5 = sorted([m for m in movs if m["credito"] > 0],
                      key=lambda m: m["credito"], reverse=True)[:5]
        if not top5:
            return "ℹ️ No encontré ingresos en el período consultado."
        periodo_txt = f" de {mes_detectado[0].capitalize()}" if mes_detectado else ""
        lineas = "\n".join(
            f"💚 {m['fecha']} | {m['concepto'][:30]} | {fmt_cop(m['credito'])}"
            for m in top5
        )
        return f"💚 *Top 5 Ingresos{periodo_txt}*\n\n{lineas}"

    # 5. Egresos
    if any(k in msg for k in ["egreso", "gasto", "salió", "salio", "pagué", "pague",
                               "gasté", "gaste", "cuánto", "cuanto"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        movs = sesion.mov_banco
        if mes_detectado:
            _, num_mes = mes_detectado
            movs = [m for m in movs if re.match(rf"\d{{4}}-{num_mes:02d}-", m.get("fecha", ""))]
        top5 = sorted([m for m in movs if m["debito"] > 0],
                      key=lambda m: m["debito"], reverse=True)[:5]
        if not top5:
            return "ℹ️ No encontré egresos en el período consultado."
        periodo_txt = f" de {mes_detectado[0].capitalize()}" if mes_detectado else ""
        lineas = "\n".join(
            f"🔴 {m['fecha']} | {m['concepto'][:30]} | {fmt_cop(m['debito'])}"
            for m in top5
        )
        return f"🔴 *Top 5 Egresos{periodo_txt}*\n\n{lineas}"

    # 6. Movimientos inusuales
    if any(k in msg for k in ["irregular", "inusual", "raro", "anormal", "sospechoso", "alerta"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        egresos_vals = [m["debito"] for m in sesion.mov_banco if m["debito"] > 0]
        if not egresos_vals:
            return "ℹ️ No encontré egresos en el extracto."
        promedio = sum(egresos_vals) / len(egresos_vals)
        umbral   = promedio * 3
        inusuales = sorted(
            [m for m in sesion.mov_banco if m["debito"] > umbral],
            key=lambda m: m["debito"], reverse=True
        )
        if not inusuales:
            return f"✅ No hay movimientos inusuales.\nPromedio de egresos: {fmt_cop(promedio)}"
        lineas = "\n".join(
            f"⚠️ {m['fecha']} | {m['concepto'][:30]} | {fmt_cop(m['debito'])}"
            for m in inusuales[:5]
        )
        return (
            f"🔍 *Movimientos Inusuales* (>3x promedio)\n"
            f"Promedio egreso: {fmt_cop(promedio)}\n"
            f"Umbral: {fmt_cop(umbral)}\n\n{lineas}"
        )

    # 7. Últimos movimientos
    if any(k in msg for k in ["último", "ultimo", "reciente", "últimas", "ultimas"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        ultimos = sesion.mov_banco[-5:][::-1]
        lineas = []
        for m in ultimos:
            val   = m["credito"] if m["credito"] > 0 else -m["debito"]
            signo = "💚" if val > 0 else "🔴"
            lineas.append(f"{signo} {m['fecha']} | {m['concepto'][:30]} | {fmt_cop(abs(val))}")
        return "📋 *Últimas 5 Transacciones*\n\n" + "\n".join(lineas)

    # 8. Nómina
    if any(k in msg for k in ["nómina", "nomina", "salario", "hada"]):
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        nomina_movs = [m for m in sesion.mov_banco if "nomi" in m.get("concepto", "").lower()]
        if not nomina_movs:
            return "ℹ️ No encontré pagos de nómina (con 'NOMI' en descripción) en el extracto."
        total_nom = sum(m["debito"] for m in nomina_movs)
        lineas = "\n".join(
            f"👥 {m['fecha']} | {m['concepto'][:35]} | {fmt_cop(m['debito'])}"
            for m in nomina_movs
        )
        return f"👥 *Pagos de Nómina*\n\n{lineas}\n\n💰 Total: {fmt_cop(total_nom)}"

    # 9. Transferencias Nequi
    if "nequi" in msg:
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        nequi_movs = [m for m in sesion.mov_banco
                      if "nequi" in m.get("concepto", "").lower()]
        if not nequi_movs:
            return "ℹ️ No encontré transferencias a Nequi en el extracto."
        total_nequi = sum(m["debito"] + m["credito"] for m in nequi_movs)
        n = len(nequi_movs)
        return (
            f"📱 *Transferencias Nequi*\n\n"
            f"Cantidad: {n} transferencia{'s' if n != 1 else ''}\n"
            f"Total: {fmt_cop(total_nequi)}"
        )

    # 10. Día específico
    meses_rx = "|".join(_MESES_NOMBRES.keys())
    rx_dia = re.search(rf"(\d{{1,2}})\s+de\s+({meses_rx})", msg)
    if rx_dia:
        if not sesion or not sesion.mov_banco:
            return _MSG_SIN_EXTRACTO
        dia        = int(rx_dia.group(1))
        nombre_mes = rx_dia.group(2)
        num_mes    = _MESES_NOMBRES[nombre_mes]
        anio       = datetime.now().year
        primera    = sesion.mov_banco[0].get("fecha", "")
        if primera:
            try:
                anio = int(primera[:4])
            except ValueError:
                pass
        fecha_buscar = f"{anio}-{num_mes:02d}-{dia:02d}"
        movs_dia = [m for m in sesion.mov_banco if m.get("fecha", "") == fecha_buscar]
        if not movs_dia:
            return f"ℹ️ No encontré transacciones el {dia} de {nombre_mes}."
        lineas = []
        for m in movs_dia:
            val   = m["credito"] if m["credito"] > 0 else -m["debito"]
            signo = "💚" if val > 0 else "🔴"
            lineas.append(f"{signo} {m['concepto'][:35]} | {fmt_cop(abs(val))}")
        return f"📅 *Movimientos del {dia} de {nombre_mes.capitalize()}*\n\n" + "\n".join(lineas)

    # Sin patrón reconocido — señal para intentar IA
    if not sesion or not sesion.mov_banco:
        return _MSG_SIN_EXTRACTO

    return None


def _consultar_gemini(mensaje: str, sesion: "SesionExtracto") -> str:
    if not _gemini:
        raise RuntimeError("Gemini no configurado")
    r    = sesion.resumen
    top5 = r.get("top10", [])[:5]
    lineas_top = "\n".join(
        f"  {m['fecha']} | {m['concepto'][:35]} | "
        f"{fmt_cop(m['credito'] if m['credito'] > 0 else m['debito'])} | "
        f"{'Abono' if m['credito'] > 0 else 'Cargo'}"
        for m in top5
    )
    contexto = (
        f"EXTRACTO ({r.get('total_tx', 0)} transacciones):\n"
        f"Ingresos: {fmt_cop(r.get('ingresos', 0))} | "
        f"Egresos: {fmt_cop(r.get('egresos', 0))} | "
        f"Saldo: {fmt_cop(r.get('saldo', 0))}\n\n"
        f"TOP 5:\n{lineas_top}"
    )[:500]
    prompt = f"{_GEMINI_SYSTEM}\n\n{contexto}\n\nUsuario: {mensaje}"
    try:
        resp = _gemini.generate_content(prompt)
    except Exception as e:
        if "429" in str(e) or "quota" in str(e).lower():
            log.warning("[Gemini] 429 quota — reintentando en 20s")
            time.sleep(20)
            resp = _gemini.generate_content(prompt)
        else:
            raise
    texto = resp.text
    if len(texto) > 1500:
        texto = texto[:1450] + "...\n_(Respuesta truncada)_"
    return texto


def responder_sin_ia(mensaje: str, sesion: "SesionExtracto | None") -> str:
    # 1. Patrones (gratis, instantáneo)
    respuesta_patron = _buscar_patron(mensaje, sesion)
    if respuesta_patron is not None:
        return respuesta_patron

    # 2. IA como fallback si hay key y sesión activa
    if GEMINI_API_KEY and sesion:
        try:
            return _consultar_gemini(mensaje, sesion)
        except Exception as e:
            log.warning(f"[Gemini] fallo — {type(e).__name__}: {e}")

    # 3. Fallback sin IA
    return "No entendí. Consulta: saldo / ingresos / gastos / nómina / inusuales / ayuda"


# ─── Endpoints HTTP ───────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "agente": "Auxiliar Contable Metrika Group", "version": "1.0"}

@app.get("/descargar/{filename}")
def descargar_archivo(filename: str):
    if ".." in filename or "/" in filename:
        raise HTTPException(400, "Nombre de archivo inválido")
    path = f"/tmp/{filename}"
    if not os.path.exists(path):
        log.error(f"Archivo no encontrado en disco: {path}")
        raise HTTPException(404, f"Archivo no encontrado: {filename}")
    size = os.path.getsize(path)
    log.info(f"Sirviendo archivo: {path} — {size:,} bytes")
    return FileResponse(
        path=path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.post("/procesar-pdf")
async def endpoint_procesar_pdf(pdf: UploadFile = File(...)):
    try:
        pdf_bytes = await pdf.read()
        log.info(f"PDF recibido por web — {len(pdf_bytes):,} bytes, filename: {pdf.filename}")
        mov_banco = await parsear_extracto_bancolombia(pdf_bytes)
        if not mov_banco:
            raise HTTPException(422, "No se pudieron extraer movimientos del PDF")
        resultado = conciliar(mov_banco)
        excel_bytes = generar_excel_conciliacion(resultado)
        filename = f"Conciliacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        url_excel = subir_excel_twilio(excel_bytes, filename)
        movimientos_banco = [
            {
                "id": f"t{i+1}",
                "fecha": m["fecha"],
                "descripcion": m["concepto"],
                "doc": m.get("doc", ""),
                "valor": m["credito"] if m["credito"] > 0 else m["debito"],
                "tipo": "CR" if m["credito"] > 0 else "DB",
                "cuenta_puc": m.get("cuenta_puc"),
                "aprobado": False,
            }
            for i, m in enumerate(resultado["mov_banco"])
        ]
        return JSONResponse({
            "status": "ok",
            "movimientos": len(mov_banco),
            "resumen": resultado["resumen"],
            "excel_url": url_excel,
            "movimientos_banco": movimientos_banco,
            "conciliados": resultado["conciliados"],
            "solo_banco":  resultado["solo_banco"],
            "solo_conta":  resultado["solo_conta"],
            "nota":        resultado["nota"],
        })
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Error en /procesar-pdf: {e}", exc_info=True)
        raise HTTPException(500, f"Error procesando PDF: {str(e)}")

# ─── Webhook WhatsApp ─────────────────────────────────────────────────────────

_MSG_SIN_PDF = "⚠️ Primero envíame el PDF de tu extracto bancario para comenzar."

@app.post("/enviar-resumen-wa")
async def enviar_resumen_wa(request: Request):
    body = await request.json()
    numero = str(body.get("numero", "")).replace("whatsapp:", "").strip()
    if not numero:
        raise HTTPException(400, "El campo 'numero' es requerido (ej. +573001234567)")

    resumen   = body.get("resumen", {})
    anomalias = body.get("anomalias")

    total_ingresos = resumen.get("totalIngresos", 0)
    total_egresos  = resumen.get("totalEgresos", 0)
    balance        = resumen.get("balanceNeto", total_ingresos - total_egresos)
    n_tx           = resumen.get("nTx", 0)
    periodo        = resumen.get("periodo", "")
    titular        = resumen.get("titular", "")

    lineas_anomalias = ""
    if anomalias:
        total_a  = anomalias.get("total_alertas", 0)
        criticas = anomalias.get("criticas", 0)
        if total_a > 0:
            lineas_anomalias = f"\n⚠️ {total_a} anomalía{'s' if total_a != 1 else ''} detectada{'s' if total_a != 1 else ''} ({criticas} crítica{'s' if criticas != 1 else ''})"
        else:
            lineas_anomalias = "\n✅ Sin anomalías detectadas"

    msg = (
        f"📊 *Reporte Ejecutivo CONTA*\n"
        f"{f'Titular: {titular}' + chr(10) if titular else ''}"
        f"Período: {periodo}\n\n"
        f"💚 Ingresos: {fmt_cop(total_ingresos)}\n"
        f"🔴 Egresos:  {fmt_cop(total_egresos)}\n"
        f"📈 Balance:  {fmt_cop(balance)}\n"
        f"📋 {n_tx} transacciones analizadas"
        f"{lineas_anomalias}\n\n"
        f"_Generado por CONTA · Metrika Group_"
    )

    try:
        enviar_whatsapp(numero, msg)
        log.info(f"Resumen WA enviado a {numero}")
        return {"status": "ok", "mensaje": "Reporte enviado por WhatsApp"}
    except Exception as e:
        log.error(f"Error enviando resumen WA a {numero}: {e}")
        raise HTTPException(500, f"Error al enviar WhatsApp: {str(e)}")


@app.post("/webhook/whatsapp")
async def webhook_whatsapp(
    background_tasks: BackgroundTasks,
    From: str = Form(...),
    Body: str = Form(default=""),
    NumMedia: str = Form(default="0"),
    MediaUrl0: str = Form(default=""),
    MediaContentType0: str = Form(default=""),
):
    limpiar_sesiones_expiradas()
    numero  = From.replace("whatsapp:", "")
    mensaje = Body.strip().lower()
    log.info(f"WA {numero}: '{Body[:80]}'")

    # ── PDF adjunto → detectar extracto o factura ────────────────────────────
    if int(NumMedia) > 0 and "pdf" in MediaContentType0.lower():
        log.info(f"WA {numero}: rama=PDF")
        background_tasks.add_task(procesar_extracto_pdf, numero, MediaUrl0)
        enviar_whatsapp(numero,
            "📄 Recibí tu PDF. Procesando... en unos segundos te envío el resultado.")
        return {"status": "procesando"}

    sesion: SesionExtracto | None = SESIONES.get(numero)

    # ── Comando: clasificar ──────────────────────────────────────────────────
    if "clasificar" in mensaje:
        log.info(f"WA {numero}: rama=clasificar")
        if not sesion or not sesion.mov_banco:
            enviar_whatsapp(numero, _MSG_SIN_PDF)
            return {"status": "ok"}
        top10 = sesion.resumen["top10"][:10]
        lineas = []
        for m in top10:
            val   = m["credito"] if m["credito"] > 0 else -m["debito"]
            puc   = m.get("cuenta_puc") or "—"
            signo = "✅" if m["credito"] > 0 else "🔴"
            lineas.append(f"{signo} {m['fecha']} | {m['concepto'][:30]} | {fmt_cop(abs(val))} | PUC {puc}")
        explicacion = (
            "ℹ️ El comando *clasificar* muestra las 10 transacciones de mayor valor "
            "con su cuenta PUC asignada. Esto te ayuda a verificar la categorización "
            "contable de tus movimientos bancarios.\n\n"
        )
        msg = explicacion + "📊 *Top 10 Transacciones con PUC*\n\n" + "\n".join(lineas)
        enviar_whatsapp(numero, msg)
        return {"status": "ok"}

    # ── Comando: pagos pendientes ────────────────────────────────────────────
    if any(k in mensaje for k in ["pagos", "pendientes", "vencimientos", "proveedores"]):
        log.info(f"WA {numero}: rama=pagos_pendientes")
        if not sesion or not sesion.mov_banco:
            enviar_whatsapp(numero, _MSG_SIN_PDF)
            return {"status": "ok"}
        background_tasks.add_task(reporte_pagos_pendientes, numero)
        return {"status": "ok"}

    # ── Comando: impuestos / DIAN ────────────────────────────────────────────
    if any(k in mensaje for k in ["impuestos", "fiscal", "tributario", "dian", "obligaciones"]):
        log.info(f"WA {numero}: rama=fiscal")
        background_tasks.add_task(reporte_fiscal, numero)
        return {"status": "ok"}

    # ── Comando: facturas ────────────────────────────────────────────────────
    if "facturas" in mensaje:
        log.info(f"WA {numero}: rama=facturas")
        if not sesion or not sesion.facturas:
            enviar_whatsapp(numero, "ℹ️ No tienes facturas registradas en esta sesión.")
            return {"status": "ok"}
        total_cp = sum(f.get("total", 0) for f in sesion.facturas)
        lineas = []
        for i, f in enumerate(sesion.facturas, 1):
            lineas.append(
                f"{i}. {f.get('emisor', 'Desconocido')[:25]}\n"
                f"   NIT: {f.get('nit', '—')} | Fecha: {f.get('fecha', '—')}\n"
                f"   Total: {fmt_cop(f.get('total', 0))}"
            )
        msg = (
            "🧾 *Facturas Registradas*\n\n"
            + "\n\n".join(lineas)
            + f"\n\n💰 *Total cuentas por pagar: {fmt_cop(total_cp)}*"
        )
        enviar_whatsapp(numero, msg)
        return {"status": "ok"}

    # ── Consulta libre → responder sin IA ───────────────────────────────────
    log.info(f"WA {numero}: rama=responder_sin_ia — '{Body[:60]}'")
    respuesta = responder_sin_ia(Body, sesion)
    enviar_whatsapp(numero, respuesta)
    return {"status": "ok"}

# ─── Tareas en background ─────────────────────────────────────────────────────

async def procesar_extracto_pdf(numero: str, media_url: str):
    try:
        log.info(f"Descargando PDF — URL: {media_url}")
        async with httpx.AsyncClient(follow_redirects=True) as client:
            resp = await client.get(
                media_url,
                auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
                follow_redirects=True,
                timeout=30.0,
            )
            if resp.status_code in (301, 302, 307, 308):
                resp = await client.get(
                    str(resp.headers.get("location", media_url)),
                    follow_redirects=True,
                    timeout=30.0,
                )
            if resp.status_code != 200:
                log.error(f"Error HTTP al descargar PDF: {resp.status_code}")
                enviar_whatsapp(numero, f"❌ No pude descargar el archivo (HTTP {resp.status_code}). Intenta de nuevo.")
                return

        pdf_bytes = resp.content
        log.info(f"PDF descargado: {len(pdf_bytes):,} bytes")

        # Detectar si es factura o extracto bancario
        texto_pdf = _extraer_texto_pdf(pdf_bytes)
        if _es_factura(texto_pdf):
            log.info(f"PDF detectado como factura para {numero}")
            await _procesar_factura(numero, texto_pdf)
            return

        mov_banco = await parsear_extracto_bancolombia(pdf_bytes)
        if not mov_banco:
            enviar_whatsapp(numero,
                "❌ No pude leer los movimientos del PDF. "
                "Asegúrate de que sea un extracto bancario en formato texto.")
            return

        clasificados = clasificar_con_reglas_locales(mov_banco)
        resumen = computar_resumen_wa(clasificados)

        # Preservar facturas de la sesión anterior si existe
        sesion_prev = SESIONES.get(numero)
        facturas_prev = sesion_prev.facturas if sesion_prev else []

        SESIONES[numero] = SesionExtracto(
            mov_banco=mov_banco,
            resumen=resumen,
            clasificados=clasificados,
            facturas=facturas_prev,
        )
        log.info(f"Sesión guardada para {numero} — {len(mov_banco)} movimientos")

        clasificados_n   = sum(1 for m in clasificados if m.get("cuenta_puc"))
        sin_clasificar_n = len(clasificados) - clasificados_n

        enviar_whatsapp(numero,
            f"✅ *Extracto procesado*\n\n"
            f"📊 Transacciones: {resumen['total_tx']}\n"
            f"💚 Ingresos: {fmt_cop(resumen['ingresos'])}\n"
            f"🔴 Egresos: {fmt_cop(resumen['egresos'])}\n"
            f"💰 Saldo: {fmt_cop(resumen['saldo'])}\n"
            f"🏷️ Clasificados PUC: {clasificados_n} | Sin clasificar: {sin_clasificar_n}\n\n"
            f"Escribe *clasificar* para ver el top 10 con cuenta PUC\n"
            f"Escribe *saldo* para ver el resumen en cualquier momento"
        )

        try:
            resultado_conc = conciliar(mov_banco)
            excel_bytes = generar_excel_conciliacion(resultado_conc)
            filename = f"Conciliacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            url_excel = subir_excel_twilio(excel_bytes, filename)
            try:
                enviar_whatsapp_media(numero, "📎 Aquí tu Excel de análisis contable:", url_excel)
            except Exception:
                enviar_whatsapp(numero, f"📎 Descarga el Excel aquí:\n{url_excel}")
        except Exception as e:
            log.warning(f"No se pudo generar Excel: {e}")

    except Exception as e:
        log.error(f"Error en procesar_extracto_pdf: {e}", exc_info=True)
        enviar_whatsapp(numero, "❌ Ocurrió un error procesando el extracto. Intenta de nuevo.")


async def _procesar_factura(numero: str, texto_pdf: str):
    try:
        datos = _extraer_datos_factura(texto_pdf)
        log.info(f"Factura procesada para {numero}: emisor={datos['emisor']}, total={datos['total']}")

        sesion = SESIONES.get(numero)
        if sesion:
            sesion.facturas.append(datos)
        else:
            SESIONES[numero] = SesionExtracto(
                mov_banco=[],
                resumen={
                    "total_tx": 0, "ingresos": 0.0, "egresos": 0.0,
                    "saldo": 0.0, "top10": [], "por_cuenta": {},
                },
                clasificados=[],
                facturas=[datos],
            )

        msg = (
            f"🧾 *Factura detectada:* {datos['tipo_factura']}\n\n"
            f"🏢 Emisor: {datos['emisor']}\n"
            f"🔢 NIT: {datos['nit']}\n"
            f"📅 Vence: {datos['fecha']}\n"
            f"💵 Subtotal: {fmt_cop(datos['subtotal'])}\n"
            f"🧮 IVA: {fmt_cop(datos['iva'])}\n"
            f"💰 Total: {fmt_cop(datos['total'])}\n"
            f"📊 Cuenta PUC sugerida: {datos['cuenta_puc']} - {datos['nombre_puc']}\n\n"
            f"✅ Agregada a cuentas por pagar\n"
            f"Escribe *facturas* para ver todas"
        )
        enviar_whatsapp(numero, msg)
    except Exception as e:
        log.error(f"Error procesando factura para {numero}: {e}", exc_info=True)
        enviar_whatsapp(numero, "❌ Error al procesar la factura. Intenta de nuevo.")


def reporte_pagos_pendientes(numero: str):
    sesion = SESIONES.get(numero)
    if not sesion:
        enviar_whatsapp(numero, _MSG_SIN_PDF)
        return

    agrupado: dict[str, dict] = defaultdict(lambda: {"n": 0, "total": 0.0, "ultimo": ""})
    for m in sesion.clasificados:
        if m["debito"] > 0:
            clave = m["concepto"][:40].strip().upper()
            agrupado[clave]["n"] += 1
            agrupado[clave]["total"] += m["debito"]
            if m["fecha"] > agrupado[clave]["ultimo"]:
                agrupado[clave]["ultimo"] = m["fecha"]

    top = sorted(agrupado.items(), key=lambda x: x[1]["total"], reverse=True)[:8]

    if not top:
        enviar_whatsapp(numero, "ℹ️ No se encontraron egresos en el extracto cargado.")
        return

    msg = "💸 *Principales Egresos del Período*\n\n"
    for concepto, d in top:
        msg += f"• {concepto[:35]}\n  {fmt_cop(d['total'])} ({d['n']} mov) — último: {d['ultimo']}\n\n"

    msg += f"_Total egresos período: {fmt_cop(sesion.resumen['egresos'])}_\n\n"
    msg += "ℹ️ Para ver facturas y vencimientos pendientes, sincroniza con OSSADO."
    enviar_whatsapp(numero, msg)


def reporte_fiscal(numero: str):
    obligaciones = _calendario_dian_proximas()
    hoy = date.today()
    msg = f"📅 *Próximas Obligaciones Fiscales*\n_{hoy.strftime('%d/%m/%Y')}_\n\n"
    for ob in obligaciones:
        venc = datetime.strptime(ob["vencimiento"], "%Y-%m-%d").date()
        dias = (venc - hoy).days
        icono = "🔴" if dias < 5 else ("⚠️" if dias <= 15 else "✅")
        msg += f"{icono} *{ob['obligacion']}*\nPeríodo: {ob['periodo']}\nVence: {ob['vencimiento']} ({dias} días)\n\n"
    msg += "_Verifica las fechas exactas según tu NIT en el calendario DIAN._"
    enviar_whatsapp(numero, msg)
